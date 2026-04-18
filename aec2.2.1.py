#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
DASHBOARD CATNAT - RISQUE SISMIQUE RPA99 - STREAMLIT + UPLOAD
✅ Upload Excel dynamique | ✅ Carte Choroplèthe 58 Wilayas | ✅ Stats | ✅ Tableau | ✅ Simulation | ✅ Nature du Risque
Design GAM Assurances - Couleurs Vertes Corporate
"""
import streamlit as st
import openpyxl
import json
from collections import defaultdict
import streamlit.components.v1 as components
from io import BytesIO

# ==========================================================
# 1. CONFIGURATION RPA99 & COORDONNÉES GPS
# ==========================================================
ZONE_RPA = {
    "ADRAR":"0","CHLEF":"III","LAGHOUAT":"0","OUM EL BOUAGHI":"IIa","BATNA":"IIa",
    "BEJAIA":"IIa","BISKRA":"0","BECHAR":"0","BLIDA":"III","BOUIRA":"IIb",
    "TAMANRASSET":"0","TEBESSA":"0","TLEMCEN":"0","TIARET":"I","TIZI OUZOU":"IIb",
    "ALGER":"III","DJELFA":"I","JIJEL":"IIa","SETIF":"IIa","SAIDA":"I",
    "SKIKDA":"IIa","SIDI BEL ABBES":"IIa","ANNABA":"IIa","GUELMA":"IIa",
    "CONSTANTINE":"IIa","MEDEA":"IIb","MOSTAGANEM":"IIa","M'SILA":"I",
    "MASCARA":"IIa","OUARGLA":"0","ORAN":"IIa","EL BAYADH":"0","ILLIZI":"0",
    "BORDJ BOU ARRERIDJ":"IIa","BOUMERDES":"III","EL TARF":"IIa","TINDOUF":"0",
    "TISSEMSILT":"IIa","EL OUED":"0","KHENCHELA":"0","SOUK AHRAS":"0",
    "TIPAZA":"III","MILA":"IIa","AIN DEFLA":"IIa","NAAMA":"0",
    "AIN TEMOUCHENT":"IIa","GHARDAIA":"0","RELIZANE":"IIa"
}

COLORS = {"0":"#10b981","I":"#059669","IIa":"#f59e0b","IIb":"#ef4444","III":"#dc2626"}
TYPE_COLORS = {"Immobilier":"#10b981","Commercial":"#059669","Industriel":"#f59e0b","Autre":"#64748b"}

GPS_COORDS = {
    "ADRAR":[27.87,-0.28],"CHLEF":[36.17,1.33],"LAGHOUAT":[33.80,2.88],
    "OUM EL BOUAGHI":[35.88,7.12],"BATNA":[35.56,6.17],"BEJAIA":[36.75,5.05],
    "BISKRA":[34.85,5.73],"BECHAR":[31.62,-2.22],"BLIDA":[36.47,2.83],
    "BOUIRA":[36.37,3.90],"TAMANRASSET":[22.79,5.52],"TEBESSA":[35.40,8.12],
    "TLEMCEN":[34.88,-1.32],"TIARET":[35.37,1.32],"TIZI OUZOU":[36.70,4.05],
    "ALGER":[36.75,3.05],"DJELFA":[34.67,3.25],"JIJEL":[36.82,5.77],
    "SETIF":[36.19,5.41],"SAIDA":[34.83,0.15],"SKIKDA":[36.87,6.90],
    "SIDI BEL ABBES":[35.19,-0.63],"ANNABA":[36.90,7.77],"GUELMA":[36.46,7.43],
    "CONSTANTINE":[36.37,6.61],"MEDEA":[36.27,2.75],"MOSTAGANEM":[35.93,0.09],
    "M'SILA":[35.71,4.54],"MASCARA":[35.40,0.14],"OUARGLA":[31.95,5.32],
    "ORAN":[35.70,-0.64],"EL BAYADH":[33.68,1.02],"ILLIZI":[26.50,8.48],
    "BORDJ BOU ARRERIDJ":[36.07,4.77],"BOUMERDES":[36.76,3.48],"EL TARF":[36.77,8.31],
    "TINDOUF":[27.67,-8.15],"TISSEMSILT":[35.61,1.81],"EL OUED":[33.37,6.86],
    "KHENCHELA":[35.43,7.14],"SOUK AHRAS":[36.29,7.95],"TIPAZA":[36.59,2.45],
    "MILA":[36.45,6.26],"AIN DEFLA":[36.25,1.97],"NAAMA":[33.27,-0.30],
    "AIN TEMOUCHENT":[35.30,-1.14],"GHARDAIA":[32.49,3.67],"RELIZANE":[35.74,0.56]
}

# ==========================================================
# 2. LECTURE EXCEL (depuis fichier uploadé)
# ==========================================================
def load_data(file_bytes):
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    
    try:
        idx_wil = headers.index('WILAYA')
        idx_cap = headers.index('CAPITAL_ASSURE')
    except ValueError:
        raise ValueError("Colonnes 'WILAYA' ou 'CAPITAL_ASSURE' introuvables dans l'Excel.")

    idx_com = headers.index('COMMUNE') if 'COMMUNE' in headers else -1
    idx_type = headers.index('TYPE') if 'TYPE' in headers else -1
    idx_prime = headers.index('PRIME_NETTE') if 'PRIME_NETTE' in headers else -1
        
    stats = defaultdict(lambda: {"cap":0.0,"zone":"","lat":0,"lng":0,"cnt":0,"prime":0.0,"types":defaultdict(int)})
    rows = []
    total_cap = 0.0
    total_prime = 0.0
    rates = {"0":0.01,"I":0.05,"IIa":0.15,"IIb":0.20,"III":0.30}
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[idx_wil]: continue
        try:
            wil = str(row[idx_wil]).split('-')[-1].strip().upper()
            cap_str = str(row[idx_cap]).replace(',','.').replace(' ','')
            cap = float(cap_str) if cap_str else 0
            if cap <= 0: continue
            
            zone = ZONE_RPA.get(wil, "IIa")
            coords = GPS_COORDS.get(wil, [36.0, 3.0])
            com = str(row[idx_com]).split('-')[-1].strip() if idx_com!=-1 and row[idx_com] else ""
            typ = str(row[idx_type]).strip() if idx_type!=-1 and row[idx_type] else "Autre"
            
            prime = 0.0
            if idx_prime != -1 and row[idx_prime]:
                prime_str = str(row[idx_prime]).replace(',','.').replace(' ','')
                try: prime = float(prime_str)
                except: prime = 0.0
                
            pml = cap * rates.get(zone, 0.15)
            
            stats[wil]["cap"] += cap
            stats[wil]["zone"] = zone
            stats[wil]["lat"] = coords[0]
            stats[wil]["lng"] = coords[1]
            stats[wil]["cnt"] += 1
            stats[wil]["prime"] += prime
            stats[wil]["types"][typ] += 1
            total_cap += cap
            total_prime += prime
            rows.append({"wil":wil, "com":com, "type":typ, "cap":cap, "prime":prime, "zone":zone, "pml":pml})
        except Exception:
            continue
            
    return dict(stats), rows, total_cap, total_prime

# ==========================================================
# 3. GÉNÉRATION HTML (CORRIGÉ: Diagrammes Nature du Risque)
# ==========================================================
def generate_html(stats, rows, total_cap, total_prime):
    total_pml = sum(r["pml"] for r in rows)
    nb_wilayas = len(stats)
    
    zone_agg = defaultdict(lambda:{"cap":0.0,"cnt":0,"prime":0.0})
    for s in stats.values():
        z = s["zone"]
        zone_agg[z]["cap"] += s["cap"]
        zone_agg[z]["cnt"] += s["cnt"]
        zone_agg[z]["prime"] += s["prime"]
    
    # ===== Aggregation par TYPE de risque =====
    type_agg = defaultdict(lambda:{"cap":0.0,"cnt":0,"prime":0.0,"pml":0.0})
    for r in rows:
        typ = r["type"]
        typ_lower = typ.lower() if typ else ""
        if any(k in typ_lower for k in ["immobilier","habitation","logement","résidentiel","residentiel","maison","villa","appartement"]):
            typ_cat = "Immobilier"
        elif any(k in typ_lower for k in ["commercial","magasin","bureau","commerce","boutique","showroom","agence"]):
            typ_cat = "Commercial"
        elif any(k in typ_lower for k in ["industriel","usine","entrepot","entrepôt","factory","depot","stockage","production"]):
            typ_cat = "Industriel"
        else:
            typ_cat = "Autre"
        type_agg[typ_cat]["cap"] += r["cap"]
        type_agg[typ_cat]["cnt"] += 1
        type_agg[typ_cat]["prime"] += r["prime"]
        type_agg[typ_cat]["pml"] += r["pml"]
    
    # ===== Zones sûres par type =====
    safe_zones_immobilier = ["0", "I"]
    safe_zones_commercial = ["0", "I", "IIa"]
    safe_zones_industriel = ["0", "I", "IIa", "IIb"]
    
    wilayas_by_zone = defaultdict(list)
    for wil, info in stats.items():
        wilayas_by_zone[info["zone"]].append({"name": wil, "cap": info["cap"], "cnt": info["cnt"], "zone": info["zone"]})
    for zone in wilayas_by_zone:
        wilayas_by_zone[zone].sort(key=lambda x: x["cap"], reverse=True)
    
    safe_immobilier = [w for z in safe_zones_immobilier if z in wilayas_by_zone for w in wilayas_by_zone[z]]
    safe_commercial = [w for z in safe_zones_commercial if z in wilayas_by_zone for w in wilayas_by_zone[z]]
    safe_industriel = [w for z in safe_zones_industriel if z in wilayas_by_zone for w in wilayas_by_zone[z]]
    
    top = sorted(stats.items(), key=lambda x:x[1]["cap"], reverse=True)[:10]
    
    wilayas_json = []
    for wil, info in stats.items():
        if wil in GPS_COORDS:
            wilayas_json.append({"name": wil, "zone": info["zone"], "cap": info["cap"], "prime": info["prime"], "cnt": info["cnt"], "lat": info["lat"], "lng": info["lng"]})
    
    stats_for_js = {wil: {"zone": info["zone"], "cap": info["cap"], "prime": info["prime"], "cnt": info["cnt"]} for wil, info in stats.items()}

    table_rows = ""
    for r in rows[:500]:
        table_rows += f"""<tr><td>{r['wil']}</td><td>{r['com']}</td><td>{r['type']}</td>
        <td><span style="background:{COLORS[r['zone']]};color:#fff;padding:6px 12px;border-radius:6px;font-size:12px;font-weight:600">{r['zone']}</span></td>
        <td style="text-align:right;font-weight:600">{r['cap']:,.0f}</td>
        <td style="text-align:right">{r['prime']:,.2f}</td>
        <td style="text-align:right;color:#dc2626;font-weight:600">{r['pml']:,.0f}</td></tr>"""

    zone_table_rows = ""
    for z in ["0","I","IIa","IIb","III"]:
        if z in zone_agg:
            d = zone_agg[z]
            pct = (d["cap"]/total_cap*100) if total_cap>0 else 0
            pml_zone = d["cap"] * {"0":0.01,"I":0.05,"IIa":0.15,"IIb":0.20,"III":0.30}.get(z, 0.15)
            zone_table_rows += f"""<tr>
                <td><span style="background:{COLORS[z]};color:#fff;padding:8px 14px;border-radius:6px;font-weight:700;display:inline-block">Zone {z}</span></td>
                <td>{sum(1 for s in stats.values() if s["zone"]==z)}</td>
                <td>{d["cnt"]:,}</td>
                <td style="text-align:right;font-weight:600">{d["cap"]:,.0f}</td>
                <td style="text-align:right">{d["prime"]:,.0f}</td>
                <td style="text-align:right;color:#dc2626;font-weight:600">{pml_zone:,.0f}</td>
                <td style="text-align:right;font-weight:600;color:#047857">{pct:.1f}%</td>
            </tr>"""

    # ===== Tableau Nature du Risque =====
    nature_table_rows = ""
    for typ_cat in ["Immobilier","Commercial","Industriel","Autre"]:
        if typ_cat in type_agg:
            d = type_agg[typ_cat]
            pct = (d["cap"]/total_cap*100) if total_cap>0 else 0
            nature_table_rows += f"""<tr>
<td><span style="background:{TYPE_COLORS[typ_cat]};color:#fff;padding:6px 12px;border-radius:6px;font-weight:600">{typ_cat}</span></td>
<td style="text-align:right;font-weight:600">{d["cnt"]:,}</td>
<td style="text-align:right;font-weight:600">{d["cap"]:,.0f}</td>
<td style="text-align:right">{d["prime"]:,.0f}</td>
<td style="text-align:right;color:#dc2626;font-weight:600">{d["pml"]:,.0f}</td>
<td style="text-align:right;font-weight:600;color:#047857">{pct:.1f}%</td>
</tr>"""

    # ===== Tableaux Zones Sûres =====
    def build_safe_table(wilayas_list):
        html = ""
        for w in wilayas_list[:12]:
            html += f"""<tr>
<td><strong>{w["name"]}</strong></td>
<td><span style="background:{COLORS[w["zone"]]};color:#fff;padding:4px 10px;border-radius:6px;font-size:12px;font-weight:600">Zone {w["zone"]}</span></td>
<td style="text-align:right;font-weight:600">{w["cap"]:,.0f}</td>
<td style="text-align:right">{w["cnt"]}</td>
<td><span style="color:#10b981;font-weight:700">✓ Sûr</span></td>
</tr>"""
        return html if html else "<tr><td colspan='5' style='text-align:center;color:#64748b'>Aucune donnée</td></tr>"
    
    safe_imm_html = build_safe_table(safe_immobilier)
    safe_com_html = build_safe_table(safe_commercial)
    safe_ind_html = build_safe_table(safe_industriel)
    
    zone3_wilayas = [wil for wil, info in stats.items() if info["zone"] == "III"]
    zone3_list = ', '.join(sorted(zone3_wilayas)) if zone3_wilayas else 'Aucune dans votre portefeuille'
    
    # ===== KPI Cards HTML for Nature Tab =====
    nature_kpi_cards = ""
    for typ_cat, color in TYPE_COLORS.items():
        if typ_cat in type_agg:
            d = type_agg[typ_cat]
            pct_cap = (d["cap"]/total_cap*100) if total_cap>0 else 0
            nature_kpi_cards += f"""<div class="nature-kpi" style="border-left-color:{color}">
<h4>{typ_cat}</h4>
<p>{d["cnt"]:,}</p>
<p style="font-size:14px;color:#64748b;margin-top:5px">{pct_cap:.1f}% du capital</p>
</div>"""

    js_code = f"""
    var GPS_COORDS = {json.dumps(GPS_COORDS)};
    
    function showTab(tabName) {{
        document.querySelectorAll('.page').forEach(p => {{
            p.classList.remove('active');
            p.style.display = 'none';
        }});
        document.querySelectorAll('.tab').forEach(b => b.classList.remove('active'));
        document.getElementById(tabName).style.display = 'block';
        setTimeout(() => document.getElementById(tabName).classList.add('active'), 10);
        if(event && event.target) {{
            event.target.classList.add('active');
        }}
        if(tabName === 'page-carte' && typeof map !== 'undefined') {{
            setTimeout(() => {{ map.invalidateSize(); }}, 100);
        }}
        // Fix for charts when switching tabs
        if(tabName === 'charts' || tabName === 'nature') {{
            setTimeout(() => {{
                window.dispatchEvent(new Event('resize'));
            }}, 200);
        }}
    }}
    
    function filt(v) {{
        const rows = document.getElementById('tb').querySelectorAll('tr');
        const val = v.toLowerCase();
        rows.forEach(r => r.style.display = r.textContent.toLowerCase().includes(val) ? '' : 'none');
    }}
    
    window.runSimulation = function() {{
        var wilaya = document.getElementById('sim-wilaya').value;
        var magnitude = parseFloat(document.getElementById('sim-magnitude').value);
        var radius = parseFloat(document.getElementById('sim-radius').value);
        var coords = GPS_COORDS[wilaya] || [36.0, 3.0];
        var affectedWilayas = [];
        var totalLoss = 0;
        var totalExposed = 0;
        var wilayas = {json.dumps(wilayas_json)};
        
        wilayas.forEach(function(w) {{
            var dist = Math.sqrt(Math.pow(w.lat - coords[0], 2) + Math.pow(w.lng - coords[1], 2)) * 111;
            if(dist <= radius) {{
                var rate = (magnitude / 10) * (1 - dist/radius) * 0.3;
                var loss = w.cap * rate;
                affectedWilayas.push({{name: w.name, capital: w.cap, loss: loss, dist: dist}});
                totalLoss += loss;
                totalExposed += w.cap;
            }}
        }});
        affectedWilayas.sort(function(a,b) {{ return b.loss - a.loss; }});
        document.getElementById('sim-result-exposed').textContent = totalExposed.toLocaleString('fr-FR') + ' DA';
        document.getElementById('sim-result-loss').textContent = totalLoss.toLocaleString('fr-FR') + ' DA';
        document.getElementById('sim-result-ratio').textContent = ((totalLoss/totalExposed)*100).toFixed(2) + '%';
        document.getElementById('sim-result-wilayas').textContent = affectedWilayas.length;
        var tbody = document.getElementById('sim-tbody');
        tbody.innerHTML = '';
        affectedWilayas.slice(0, 15).forEach(function(w) {{
            tbody.innerHTML += '<tr><td>'+w.name+'</td><td>'+w.dist.toFixed(1)+' km</td>' +
                '<td style="text-align:right;font-weight:600">'+w.capital.toLocaleString('fr-FR')+'</td>' +
                '<td style="text-align:right;color:#dc2626;font-weight:700">'+w.loss.toLocaleString('fr-FR')+' DA</td></tr>';
        }});
        document.getElementById('sim-results').style.display = 'block';
        document.getElementById('sim-results').scrollIntoView({{behavior: 'smooth'}});
    }};
    
    setTimeout(function() {{
        var fmt = v => v.toLocaleString('fr-FR');
        var animate = (id, end, suffix) => {{
            var el = document.getElementById(id); 
            if(!el) return;
            var st = Date.now();
            (function run() {{
                var p = Math.min((Date.now()-st)/1500, 1);
                el.textContent = fmt(Math.floor(end*p)) + suffix;
                if(p < 1) requestAnimationFrame(run);
            }})();
        }};
        animate('kpi-cap', {total_cap}, ' DA');
        animate('kpi-prime', {total_prime}, ' DA');
        animate('kpi-pml', {total_pml}, ' DA');
        var elWil = document.getElementById('kpi-wil');
        if(elWil) elWil.textContent = '{nb_wilayas}';
        var elCtr = document.getElementById('kpi-ctr');
        if(elCtr) elCtr.textContent = '{len(rows):,}';
    }}, 300);
    
    // ===== CARTE CHOROPLETHE 58 WILAYAS =====
    var map = L.map('map').setView([28.0, 3.0], 5);
    L.tileLayer('https://{{s}}.basemaps.cartocdn.com/rastertiles/voyager/{{z}}/{{x}}/{{y}}{{r}}.png', {{
        attribution: '© OpenStreetMap © CARTO', subdomains: 'abcd', maxZoom: 10
    }}).addTo(map);
    
    var colors = {json.dumps(COLORS)};
    var zoneRPA = {json.dumps(ZONE_RPA)};
    var statsData = {json.dumps(stats_for_js)};
    
    function getZoneStyle(feature) {{
        var wilayaName = feature.properties.name ? feature.properties.name.toUpperCase().replace(/'/g, "''") : "";
        var zone = zoneRPA[wilayaName] || "IIa";
        return {{
            fillColor: colors[zone],
            weight: 1.5,
            opacity: 1,
            color: '#ffffff',
            fillOpacity: 0.85
        }};
    }}
    
    function fmt(v) {{ return v.toLocaleString('fr-FR'); }}
    
    fetch('https://raw.githubusercontent.com/fr33dz/Algeria-geojson/master/all-wilayas.geojson')
        .then(response => {{
            if(!response.ok) throw new Error('GeoJSON non chargé');
            return response.json();
        }})
        .then(data => {{
            L.geoJSON(data, {{
                style: getZoneStyle,
                onEachFeature: function(feature, layer) {{
                    var name = feature.properties.name || "Inconnue";
                    var nameUpper = name.toUpperCase().replace(/'/g, "''");
                    var info = statsData[nameUpper] || {{zone: zoneRPA[nameUpper] || "IIa", cap: 0, prime: 0, cnt: 0}};
                    var zone = info.zone;
                    layer.bindPopup(
                        '<b>' + name + '</b><br>' +
                        'Zone RPA99: <span style="background:'+colors[zone]+';color:#fff;padding:3px 8px;border-radius:4px;font-weight:600">'+zone+'</span><br>' +
                        'Capital: ' + fmt(info.cap) + ' DA<br>' +
                        'Contrats: ' + info.cnt + '<br>' +
                        'Prime: ' + fmt(info.prime) + ' DA'
                    );
                    layer.on('mouseover', function(e) {{ this.setStyle({{weight: 3, color: '#10b981'}}); }});
                    layer.on('mouseout', function(e) {{ this.setStyle(getZoneStyle(this.feature)); }});
                }}
            }}).addTo(map);
        }})
        .catch(err => {{
            console.log('Fallback cercles GPS:', err);
            var wilayas = {json.dumps(wilayas_json)};
            wilayas.forEach(w => {{
                L.circleMarker([w.lat, w.lng], {{
                    radius: Math.max(8, Math.min(20, Math.sqrt(w.cap)/15000)),
                    fillColor: colors[w.zone], color: '#FFF', weight: 2, fillOpacity: 0.85
                }}).addTo(map).bindPopup('<b>'+w.name+'</b><br>Zone: '+w.zone+'<br>Capital: '+fmt(w.cap)+' DA');
            }});
        }});
    
    var legend = L.control({{position: 'bottomright'}});
    legend.onAdd = function(map) {{
        var div = L.DomUtil.create('div', 'info legend');
        div.style.cssText = 'background:white;padding:12px;border-radius:8px;box-shadow:0 2px 8px rgba(0,0,0,0.15);font-size:12px;line-height:1.8';
        div.innerHTML = '<b>Zones RPA99</b><br>';
        ["0","I","IIa","IIb","III"].forEach(z => {{
            div.innerHTML += '<i style="background:'+colors[z]+';width:18px;height:18px;display:inline-block;margin-right:6px;border-radius:4px;border:1px solid #fff"></i> Zone ' + z + '<br>';
        }});
        return div;
    }};
    legend.addTo(map);
    
    // ===== GRAPHIQUES CHART.JS =====
    new Chart(document.getElementById('c1'), {{
        type: 'doughnut',
        data: {{
            labels: {json.dumps([f"Zone {z}" for z in zone_agg.keys()])},
            datasets: [{{ data: {json.dumps([d["cap"] for d in zone_agg.values()])}, backgroundColor: {json.dumps([COLORS[z] for z in zone_agg.keys()])}, borderWidth: 0 }}]
        }},
        options: {{ responsive: true, maintainAspectRatio: false, cutout: '70%', plugins: {{ legend: {{ position: 'bottom' }} }} }}
    }});
    
    new Chart(document.getElementById('c2'), {{
        type: 'bar',
        data: {{
            labels: {json.dumps([w[0] for w in top])},
            datasets: [{{ label: 'Capital (M DA)', data: {json.dumps([w[1]["cap"]/1e6 for w in top])}, backgroundColor: '#047857', borderRadius: 8 }}]
        }},
        options: {{ responsive: true, maintainAspectRatio: false, indexAxis: 'y' }}
    }});
    
    // CORRECTED: Nature Tab Charts (Unique IDs)
    new Chart(document.getElementById('c-nature-1'), {{
        type: 'doughnut',
        data: {{
            labels: {json.dumps(list(type_agg.keys()))},
            datasets: [{{ data: {json.dumps([type_agg[t]["cnt"] for t in type_agg.keys()])}, backgroundColor: {json.dumps([TYPE_COLORS[t] for t in type_agg.keys()])}, borderWidth: 0 }}]
        }},
        options: {{ responsive: true, maintainAspectRatio: false, cutout: '70%', plugins: {{ legend: {{ position: 'bottom' }} }} }}
    }});
    
    new Chart(document.getElementById('c-nature-2'), {{
        type: 'bar',
        data: {{
            labels: {json.dumps(list(type_agg.keys()))},
            datasets: [{{ label: 'Capital (M DA)', data: {json.dumps([type_agg[t]["cap"]/1e6 for t in type_agg.keys()])}, backgroundColor: {json.dumps([TYPE_COLORS[t] for t in type_agg.keys()])}, borderRadius: 8 }}]
        }},
        options: {{ responsive: true, maintainAspectRatio: false }}
    }});
    """

    html = f"""<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>GAM Assurances - Dashboard Risques Sismiques RPA99</title>
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css" />
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:'Segoe UI',Tahoma,Geneva,Verdana,sans-serif;background:#f8fafc;color:#1e293b;line-height:1.6}}

header{{
    background:linear-gradient(135deg,#047857 0%,#059669 100%);
    box-shadow:0 4px 12px rgba(0,0,0,0.1);
    position:relative;
    overflow:hidden;
}}
header::before{{
    content:'';
    position:absolute;
    top:0;
    left:0;
    right:0;
    height:4px;
    background:linear-gradient(90deg,#10b981,#059669,#f59e0b);
}}
.header-content{{
    max-width:1400px;
    margin:0 auto;
    padding:30px 40px;
    position:relative;
    z-index:1;
}}
.brand{{
    display:flex;
    align-items:center;
    gap:15px;
    margin-bottom:20px;
}}
.brand-logo{{
    width:60px;
    height:60px;
    background:white;
    border-radius:12px;
    display:flex;
    align-items:center;
    justify-content:center;
    font-size:28px;
    box-shadow:0 4px 12px rgba(0,0,0,0.15);
}}
.brand-text h1{{
    color:white;
    font-size:28px;
    font-weight:700;
    margin-bottom:5px;
    letter-spacing:-0.5px;
}}
.brand-text p{{
    color:rgba(255,255,255,0.9);
    font-size:14px;
    font-weight:500;
}}

.nav-container{{
    background:rgba(255,255,255,0.1);
    backdrop-filter:blur(10px);
    border-radius:12px;
    padding:8px;
    display:inline-flex;
    gap:8px;
    flex-wrap:wrap;
}}
.tab{{
    padding:12px 24px;
    background:transparent;
    border:none;
    color:rgba(255,255,255,0.85);
    cursor:pointer;
    border-radius:8px;
    font-weight:600;
    font-size:14px;
    transition:all 0.3s;
    white-space:nowrap;
}}
.tab:hover{{
    background:rgba(255,255,255,0.15);
    color:white;
}}
.tab.active{{
    background:white;
    color:#047857;
    box-shadow:0 4px 12px rgba(0,0,0,0.1);
}}

.container{{
    max-width:1400px;
    margin:0 auto;
    padding:40px;
}}
.page{{
    display:none;
    animation:fadeIn 0.5s ease;
}}
.page.active{{
    display:block;
}}
@keyframes fadeIn{{
    from{{opacity:0;transform:translateY(20px)}}
    to{{opacity:1;transform:translateY(0)}}
}}

.section-title{{
    font-size:24px;
    font-weight:700;
    color:#1e293b;
    margin-bottom:30px;
    padding-bottom:15px;
    border-bottom:3px solid #10b981;
    display:inline-block;
}}

.kpi-grid{{
    display:grid;
    grid-template-columns:repeat(auto-fit,minmax(280px,1fr));
    gap:24px;
    margin-bottom:40px;
}}
.kpi{{
    background:white;
    padding:30px;
    border-radius:16px;
    box-shadow:0 4px 16px rgba(0,0,0,0.08);
    border-top:4px solid #10b981;
    transition:all 0.3s;
    position:relative;
    overflow:hidden;
}}
.kpi::before{{
    content:'';
    position:absolute;
    top:0;
    right:0;
    width:120px;
    height:120px;
    background:linear-gradient(135deg,rgba(16,185,129,0.05) 0%,rgba(16,185,129,0) 100%);
    border-radius:0 0 0 100%;
}}
.kpi:hover{{
    transform:translateY(-8px);
    box-shadow:0 12px 32px rgba(0,0,0,0.12);
}}
.kpi h3{{
    font-size:13px;
    color:#64748b;
    text-transform:uppercase;
    letter-spacing:1px;
    font-weight:600;
    margin-bottom:12px;
    display:flex;
    align-items:center;
    gap:8px;
}}
.kpi p{{
    font-size:32px;
    font-weight:700;
    color:#047857;
    line-height:1.2;
}}

.table-container{{
    background:white;
    border-radius:16px;
    overflow:hidden;
    box-shadow:0 4px 16px rgba(0,0,0,0.08);
    margin-bottom:30px;
}}
.table-header{{
    background:linear-gradient(135deg,#047857 0%,#059669 100%);
    color:white;
    padding:20px 30px;
    font-size:18px;
    font-weight:700;
}}
.search-box{{
    padding:20px 30px;
    background:#f8fafc;
    border-bottom:1px solid #e2e8f0;
}}
.search-box input{{
    width:100%;
    max-width:500px;
    padding:12px 20px;
    border:2px solid #e2e8f0;
    border-radius:10px;
    font-size:14px;
    transition:all 0.3s;
}}
.search-box input:focus{{
    outline:none;
    border-color:#10b981;
    box-shadow:0 0 0 4px rgba(16,185,129,0.1);
}}
.table-wrap{{
    max-height:600px;
    overflow-y:auto;
}}
table{{
    width:100%;
    border-collapse:collapse;
    font-size:14px;
}}
th{{
    background:#f1f5f9;
    color:#475569;
    font-weight:700;
    text-transform:uppercase;
    font-size:12px;
    letter-spacing:0.5px;
    padding:16px 20px;
    text-align:left;
    position:sticky;
    top:0;
    z-index:10;
}}
td{{
    padding:16px 20px;
    border-bottom:1px solid #f1f5f9;
    color:#475569;
}}
tr:hover{{
    background:#f8fafc;
}}

.sim-panel{{
    background:white;
    border-radius:16px;
    padding:40px;
    box-shadow:0 4px 16px rgba(0,0,0,0.08);
    margin-bottom:30px;
}}
.sim-header{{
    background:linear-gradient(135deg,#dc2626 0%,#ef4444 100%);
    color:white;
    padding:25px 30px;
    border-radius:12px;
    margin-bottom:30px;
}}
.sim-header h2{{
    font-size:24px;
    font-weight:700;
    display:flex;
    align-items:center;
    gap:12px;
}}
.form-grid{{
    display:grid;
    grid-template-columns:repeat(auto-fit,minmax(280px,1fr));
    gap:24px;
    margin-bottom:30px;
}}
.form-group{{
    display:flex;
    flex-direction:column;
    gap:8px;
}}
.form-group label{{
    font-weight:600;
    color:#475569;
    font-size:14px;
}}
.form-group input,.form-group select{{
    padding:14px 18px;
    border:2px solid #e2e8f0;
    border-radius:10px;
    font-size:15px;
    transition:all 0.3s;
    background:white;
}}
.form-group input:focus,.form-group select:focus{{
    outline:none;
    border-color:#10b981;
    box-shadow:0 0 0 4px rgba(16,185,129,0.1);
}}
.btn-sim{{
    background:linear-gradient(135deg,#dc2626 0%,#ef4444 100%);
    color:white;
    border:none;
    padding:16px 40px;
    border-radius:10px;
    font-size:16px;
    font-weight:700;
    cursor:pointer;
    transition:all 0.3s;
    box-shadow:0 4px 16px rgba(220,38,38,0.3);
}}
.btn-sim:hover{{
    transform:translateY(-2px);
    box-shadow:0 8px 24px rgba(220,38,38,0.4);
}}
.sim-results{{
    display:none;
    background:#fef2f2;
    border:2px solid #fecaca;
    border-radius:12px;
    padding:30px;
    margin-top:30px;
}}
.sim-results h3{{
    color:#dc2626;
    font-size:20px;
    font-weight:700;
    margin-bottom:25px;
}}
.sim-kpis{{
    display:grid;
    grid-template-columns:repeat(auto-fit,minmax(220px,1fr));
    gap:20px;
    margin-bottom:30px;
}}
.sim-kpi{{
    background:white;
    padding:24px;
    border-radius:12px;
    text-align:center;
    box-shadow:0 4px 12px rgba(0,0,0,0.08);
    border-left:4px solid #dc2626;
}}
.sim-kpi h4{{
    color:#64748b;
    font-size:12px;
    text-transform:uppercase;
    letter-spacing:1px;
    margin-bottom:10px;
    font-weight:600;
}}
.sim-kpi p{{
    color:#dc2626;
    font-size:26px;
    font-weight:700;
}}

.charts-grid{{
    display:grid;
    grid-template-columns:repeat(auto-fit,minmax(450px,1fr));
    gap:24px;
}}
.chart-box{{
    background:white;
    padding:30px;
    border-radius:16px;
    box-shadow:0 4px 16px rgba(0,0,0,0.08);
    min-height: 350px; /* FIX: Ensure chart container has height */
}}

#map{{
    height:650px;
    border-radius:16px;
    box-shadow:0 8px 24px rgba(0,0,0,0.12);
    border:2px solid #e2e8f0;
}}

/* Nature du Risque Styles */
.nature-kpi-grid{{
    display:grid;
    grid-template-columns:repeat(auto-fit,minmax(220px,1fr));
    gap:20px;
    margin-bottom:30px;
}}
.nature-kpi{{
    background:white;
    padding:25px;
    border-radius:12px;
    box-shadow:0 4px 12px rgba(0,0,0,0.08);
    border-left:4px solid #10b981;
    text-align:center;
}}
.nature-kpi h4{{
    color:#64748b;
    font-size:13px;
    text-transform:uppercase;
    margin-bottom:10px;
}}
.nature-kpi p{{
    font-size:28px;
    font-weight:700;
    color:#047857;
}}

.safe-zone-box{{
    background:white;
    border-radius:12px;
    padding:25px;
    box-shadow:0 4px 12px rgba(0,0,0,0.08);
    margin-bottom:25px;
    border-top:4px solid #10b981;
}}
.safe-zone-box.commercial{{border-top-color:#059669}}
.safe-zone-box.industriel{{border-top-color:#f59e0b}}
.safe-zone-box h4{{
    color:#1e293b;
    font-size:18px;
    font-weight:700;
    margin-bottom:15px;
    display:flex;
    align-items:center;
    gap:10px;
}}
.safe-zone-box p{{
    color:#64748b;
    font-size:14px;
    margin-bottom:15px;
}}

.alert-box{{
    background:#fef2f2;
    border:2px solid #fecaca;
    border-radius:12px;
    padding:20px 25px;
    margin-top:20px;
}}
.alert-box h4{{
    color:#dc2626;
    font-size:16px;
    font-weight:700;
    margin-bottom:10px;
}}
.alert-box p{{
    color:#991b1b;
    font-size:14px;
}}

@media(max-width:900px){{
    .header-content{{padding:25px 20px}}
    .brand-text h1{{font-size:22px}}
    .container{{padding:25px 20px}}
    .charts-grid{{grid-template-columns:1fr}}
    .kpi-grid{{grid-template-columns:1fr}}
    .sim-panel{{padding:25px 20px}}
}}
</style>
</head>
<body>
<header>
<div class="header-content">
<div class="brand">
<div class="brand-text">
<h1>🗺️ GAM Assurances</h1>
<h2>Votre protection est notre engagement</h2>
<h3>Dashboard Analyse des Risques Sismiques - RPA99</h3>
<p>Portefeuille CATNAT 2023-2025 • Conformité réglementaire Algérie</p>
</div>
</div>
<div class="nav-container">
<button class="tab active" onclick="showTab('page-carte')">🗺️ Carte des Risques</button>
<button class="tab" onclick="showTab('stats')">📊 Statistiques</button>
<button class="tab" onclick="showTab('table')">📋 Contrats</button>
<button class="tab" onclick="showTab('charts')">📈 Analyses</button>
<button class="tab" onclick="showTab('simulation')" style="background:#dc2626;color:white">⚡ Simulation</button>
<button class="tab" onclick="showTab('nature')" style="background:#10b981;color:white">🏢 Nature du Risque</button>
</div>
</div>
</header>

<div class="container">
<!-- Carte -->
<div id="page-carte" class="page active">
<h2 class="section-title">Cartographie des Risques (58 Wilayas)</h2>
<div class="kpi-grid">
<div class="kpi" style="border-top-color:#10b981"><h3>💰 Capital Total Assuré</h3><p id="kpi-cap">0 DA</p></div>
<div class="kpi" style="border-top-color:#059669"><h3>📊 Prime Nette Totale</h3><p id="kpi-prime">0 DA</p></div>
<div class="kpi" style="border-top-color:#dc2626"><h3>⚠️ PML Estimée</h3><p id="kpi-pml" style="color:#dc2626">0 DA</p></div>
<div class="kpi" style="border-top-color:#f59e0b"><h3>🏛️ Wilayas Actives</h3><p id="kpi-wil">0</p></div>
<div class="kpi" style="border-top-color:#8b5cf6"><h3>📋 Nombre de Contrats</h3><p id="kpi-ctr">0</p></div>
</div>
<div id="map"></div>
</div>

<!-- Statistiques -->
<div id="stats" class="page">
<h2 class="section-title">📊 Statistiques Détaillées</h2>
<div class="kpi-grid">
<div class="kpi" style="border-top-color:#10b981"><h3>Capital Total</h3><p>{total_cap:,.0f} DA</p></div>
<div class="kpi" style="border-top-color:#059669"><h3>Prime Nette</h3><p>{total_prime:,.0f} DA</p></div>
<div class="kpi" style="border-top-color:#dc2626"><h3>PML Scénario Majeur</h3><p style="color:#dc2626">{total_pml:,.0f} DA</p></div>
<div class="kpi" style="border-top-color:#f59e0b"><h3>Wilayas</h3><p>{nb_wilayas}</p></div>
</div>
<div class="table-container">
<div class="table-header">📊 Répartition par Zone Sismique RPA99</div>
<div class="table-wrap"><table><thead><tr><th>Zone</th><th>Wilayas</th><th>Contrats</th><th>Capital (DA)</th><th>Prime (DA)</th><th>PML (DA)</th><th>% Capital</th></tr></thead><tbody>{zone_table_rows}</tbody></table></div>
</div>
</div>

<!-- Tableau -->
<div id="table" class="page">
<div class="table-container">
<div class="table-header">📋 Détail des Contrats</div>
<div class="search-box"><input type="text" placeholder="🔍 Rechercher (Wilaya, Commune, Type...)" onkeyup="filt(this.value)"></div>
<div class="table-wrap"><table><thead><tr><th>Wilaya</th><th>Commune</th><th>Type</th><th>Zone</th><th>Capital</th><th>Prime</th><th>PML</th></tr></thead><tbody id="tb">{table_rows}</tbody></table></div>
</div>
</div>

<!-- Graphiques -->
<div id="charts" class="page">
<h2 class="section-title">📈 Analyses Graphiques</h2>
<div class="charts-grid">
<div class="chart-box"><canvas id="c1"></canvas></div>
<div class="chart-box"><canvas id="c2"></canvas></div>
</div>
</div>

<!-- Simulation -->
<div id="simulation" class="page">
<div class="sim-panel">
<div class="sim-header"><h2>⚡ Simulation de Catastrophe Sismique</h2><p style="margin-top:10px;opacity:0.9">Évaluez l'impact financier d'un séisme sur votre portefeuille</p></div>
<div class="form-grid">
<div class="form-group"><label>Wilaya Épicentre</label><select id="sim-wilaya">
<option value="ALGER">Alger (Zone III)</option><option value="CHLEF">Chlef (Zone III)</option><option value="BOUMERDES">Boumerdes (Zone III)</option>
<option value="BLIDA">Blida (Zone III)</option><option value="TIPAZA">Tipaza (Zone III)</option><option value="ORAN">Oran (Zone IIa)</option>
<option value="CONSTANTINE">Constantine (Zone IIa)</option><option value="SETIF">Sétif (Zone IIa)</option><option value="BEJAIA">Béjaïa (Zone IIa)</option><option value="TIZI OUZOU">Tizi Ouzou (Zone IIb)</option>
</select></div>
<div class="form-group"><label>Magnitude (Échelle Richter)</label><input type="number" id="sim-magnitude" min="4" max="9" step="0.1" value="6.5"></div>
<div class="form-group"><label>Rayon d'impact (km)</label><input type="number" id="sim-radius" min="10" max="200" value="100"></div>
</div>
<button class="btn-sim" onclick="runSimulation()">🚀 Lancer la Simulation</button>
<div id="sim-results" class="sim-results"><h3>📊 Résultats de la Simulation</h3>
<div class="sim-kpis">
<div class="sim-kpi"><h4>Capital Exposé</h4><p id="sim-result-exposed">0 DA</p></div>
<div class="sim-kpi"><h4>Pertes Estimées</h4><p id="sim-result-loss">0 DA</p></div>
<div class="sim-kpi"><h4>Ratio Pertes/Capital</h4><p id="sim-result-ratio">0%</p></div>
<div class="sim-kpi"><h4>Wilayas Touchées</h4><p id="sim-result-wilayas">0</p></div>
</div>
<div class="table-container" style="margin-top:20px"><div class="table-header" style="background:#dc2626">Détail par Wilaya Impactée</div>
<div class="table-wrap"><table><thead><tr><th>Wilaya</th><th>Distance</th><th>Capital</th><th>Perte Estimée</th></tr></thead><tbody id="sim-tbody"></tbody></table></div></div>
</div>
</div>
</div>

<!-- ===== Nature du Risque ===== -->
<div id="nature" class="page">
    <h2 class="section-title">🏢 Nature des Risques par Type de Bien</h2>

    <div class="nature-kpi-grid">
        {nature_kpi_cards}
    </div>

    <div class="charts-grid">
        <!-- CORRECTED: Unique IDs for Nature Charts -->
        <div class="chart-box"><canvas id="c-nature-1"></canvas></div>
        <div class="chart-box"><canvas id="c-nature-2"></canvas></div>
    </div>

    <h3 style="margin:35px 0 20px 0;color:#1e293b;font-size:20px;font-weight:700">📋 Répartition par Type</h3>
    <div class="table-container">
        <div class="table-header">Détail par Catégorie de Bien</div>
        <div class="table-wrap"><table><thead><tr><th>Type</th><th>Nb Contrats</th><th>Capital (DA)</th><th>Prime (DA)</th><th>PML (DA)</th><th>% Portefeuille</th></tr></thead><tbody>{nature_table_rows}</tbody></table></div>
    </div>

    <h2 style="margin:45px 0 25px 0;color:#1e293b;font-size:24px;font-weight:700;border-bottom:3px solid #10b981;padding-bottom:15px;display:inline-block">📍 Zones Recommandées par Type</h2>

    <div class="safe-zone-box">
        <h4>🏠 Immobilier / Habitation</h4>
        <p><strong>Zones recommandées:</strong> Zone 0 (Très Faible) et Zone I (Faible)<br>
        Idéal pour: logements, villas, appartements, résidences</p>
        <div class="table-wrap" style="max-height:350px"><table><thead><tr><th>Wilaya</th><th>Zone</th><th>Capital (DA)</th><th>Contrats</th><th>Sécurité</th></tr></thead><tbody>{safe_imm_html}</tbody></table></div>
    </div>

    <div class="safe-zone-box commercial">
        <h4>🏪 Commercial / Bureaux</h4>
        <p><strong>Zones recommandées:</strong> Zones 0, I et IIa (Faible à Moyenne)<br>
        Idéal pour: magasins, bureaux, commerces, agences</p>
        <div class="table-wrap" style="max-height:350px"><table><thead><tr><th>Wilaya</th><th>Zone</th><th>Capital (DA)</th><th>Contrats</th><th>Sécurité</th></tr></thead><tbody>{safe_com_html}</tbody></table></div>
    </div>

    <div class="safe-zone-box industriel">
        <h4>🏭 Industriel / Entreposage</h4>
        <p><strong>Zones recommandées:</strong> Zones 0, I, IIa et IIb (avec normes adaptées)<br>
        Idéal pour: usines, entrepôts, sites de production</p>
        <div class="table-wrap" style="max-height:350px"><table><thead><tr><th>Wilaya</th><th>Zone</th><th>Capital (DA)</th><th>Contrats</th><th>Sécurité</th></tr></thead><tbody>{safe_ind_html}</tbody></table></div>
    </div>

    <div class="alert-box">
        <h4>⚠️ Attention: Zones à Risque Élevé (Zone III)</h4>
        <p><strong>Wilayas concernées:</strong> {zone3_list}<br>
        Pour ces zones, des études géotechniques et des normes parasismiques renforcées sont obligatoires pour tous types de constructions.</p>
    </div>
</div>

</div>

<script>{js_code}</script>
</body>
</html>"""
    
    return html

# ==========================================================
# 4. STREAMLIT APP PRINCIPALE
# ==========================================================
def main():
    st.set_page_config(page_title="🗺️ GAM Assurances - RPA99", layout="wide", page_icon="🛡️")
    
    # Header
    st.markdown("""
    <style>
    .main-header {background: linear-gradient(135deg, #047857 0%, #059669 100%); padding: 20px; border-radius: 10px; color: white; margin-bottom: 20px;}
    .main-header h1 {margin: 0; font-size: 28px;}
    .main-header p {margin: 5px 0 0 0; opacity: 0.9;}
    </style>
    <div class="main-header">
        <h1>🗺️ GAM Assurances</h1>
        <p>Dashboard Analyse des Risques Sismiques - RPA99 | Portefeuille CATNAT 2023-2025</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Upload
    uploaded_file = st.file_uploader("📁 Choisir un fichier Excel", type=['xlsx', 'xls'], help="Sélectionnez votre fichier avec les colonnes: WILAYA, CAPITAL_ASSURE, COMMUNE, TYPE, PRIME_NETTE")
    
    if uploaded_file is None:
        st.info("👆 Veuillez uploader votre fichier Excel pour générer le dashboard")
        st.markdown("""
        ### 📋 Format requis :
        | Colonne | Requis | Exemple |
        |---------|--------|---------|
        | `WILAYA` | ✅ | `01-ADRAR` |
        | `CAPITAL_ASSURE` | ✅ | `15000000` |
        | `COMMUNE` | ❌ | `Adrar` |
        | `TYPE` | ❌ | `Immobilier` |
        | `PRIME_NETTE` | ❌ | `25000.50` |
        """)
        return
    
    # Process
    with st.spinner('📊 Traitement en cours...'):
        try:
            stats, rows, total_cap, total_prime = load_data(uploaded_file.getvalue())
            if not stats:
                st.error("❌ Aucune donnée valide. Vérifiez les colonnes de votre Excel.")
                return
        except Exception as e:
            st.error(f"❌ Erreur: {str(e)}")
            return
    
    st.success(f"✅ **{len(rows)} contrats** chargés depuis `{uploaded_file.name}`")
    
    # Generate & Display
    html_dashboard = generate_html(stats, rows, total_cap, total_prime)
    
    # Render in Streamlit
    st.subheader("📊 Dashboard RPA99")
    components.html(html_dashboard, height=1400, scrolling=True)
    
    st.markdown("---")
    st.caption("🔐 Dashboard généré localement • Aucune donnée envoyée vers un serveur externe • Conformité RPA99 Algérie")

if __name__ == "__main__":
    main()